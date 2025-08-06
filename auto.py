# python3 -m pip install openpyxl 


import openpyxl as xl
from openpyxl.chart import BarChart, Reference  

def process_workbook(filename):
    wb = xl.load_workbook(filename) # storing in workbook object
    sheet =wb['Sheet1']

    for row in range(2,sheet.max_row+1):    #to include the last row too
        cell = sheet.cell(row,3)
        new_price = cell.value*0.9 # reducing the price value by 10%
        new_price_cell = sheet.cell(row,4)
        new_price_cell.value=new_price

    #Values will store the information in the 4th column
    values =Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()  #creating a object
    chart.add_data(values)  #passing the values
    sheet.add_chart(chart,'f2')


    wb.save(filename)

filename =input("enter file name:")
process_workbook(filename)
